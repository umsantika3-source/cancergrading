import os
import glob
import json
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import config


_PURPLE     = "9400D3"
_PINK_LIGHT = "F3E5F5"
_PINK_MED   = "CE93D8"
_WHITE      = "FFFFFF"
_GREY       = "F5F5F5"
_DARK       = "212121"


def _versioned_path(path):
    if not os.path.exists(path):
        return path
    base, ext = os.path.splitext(path)
    n = 2
    while os.path.exists(f"{base}_{n}{ext}"):
        n += 1
    return f"{base}_{n}{ext}"


def _migrate_eval(d, path=None):
    """
    Normalises old-format eval.json dicts to the current schema in-place.

    Old format (Exp1/Exp2):
      keys: train_acc, val_acc, test_acc, train_loss, val_loss, test_loss,
            precision, recall, f1, train_time, epochs_trained, cm,
            per_class_metrics (list)
      exp / model: absent or None

    Intermediate format (Exp3/Exp4 pre-update):
      has exp, model, accuracy, confusion_matrix, macro_f1, weighted_f1
      per_class_metrics: dict but may use wrong class-name keys
      missing: macro_precision, macro_recall, test_loss, train_acc_final,
               val_acc_final, train_loss_final, val_loss_final,
               epochs_trained, training_time_seconds, best_val_acc
    """
    # ── Infer exp / model from filename if absent ────────────────────────────
    if (not d.get("exp") or not d.get("model")) and path:
        stem = os.path.basename(path).replace("_eval.json", "")
        parts = stem.split("_", 1)           # "Exp1_AlexNet" → ["Exp1", "AlexNet"]
        if len(parts) == 2:
            d["exp"]   = d.get("exp")   or parts[0]
            d["model"] = d.get("model") or parts[1]

    exp = d.get("exp", "")

    # ── Old field → new field renames ────────────────────────────────────────
    _remap = [
        ("test_acc",   "accuracy"),
        ("precision",  "macro_precision"),
        ("recall",     "macro_recall"),
        ("f1",         "macro_f1"),
        ("cm",         "confusion_matrix"),
        ("train_time", "training_time_seconds"),
        ("train_acc",  "train_acc_final"),
        ("val_acc",    "val_acc_final"),
        ("train_loss", "train_loss_final"),
        ("val_loss",   "val_loss_final"),
    ]
    for old_key, new_key in _remap:
        if old_key in d and new_key not in d:
            d[new_key] = d[old_key]

    # best_val_acc approximation from val_acc_final when missing
    if "best_val_acc" not in d:
        d["best_val_acc"] = d.get("val_acc_final", 0)

    # Infer activation / attention from experiment label
    if "activation" not in d:
        d["activation"] = "ReLU" if exp == "Exp1" else "GELU"
    if "attention" not in d:
        if exp in ("Exp3", "Exp4"):
            d["attention"] = config.ATTENTION_TYPE or "SE"
        else:
            d["attention"] = "None"

    # ── per_class_metrics: list → dict ───────────────────────────────────────
    per = d.get("per_class_metrics", {})
    if isinstance(per, list):
        new_per = {}
        for item in per:
            cls = item.get("Class", "?")
            new_per[cls] = {
                "precision": item.get("Precision", 0),
                "recall":    item.get("Recall",    0),
                "f1":        item.get("F1",         0),
            }
        d["per_class_metrics"] = new_per

    # Ensure zeros for any still-missing numeric fields
    for field in ("macro_precision", "macro_recall", "test_loss",
                  "train_acc_final", "val_acc_final",
                  "train_loss_final", "val_loss_final",
                  "epochs_trained", "training_time_seconds"):
        d.setdefault(field, 0)

    return d


def _hdr(cell, text, bold=True, size=10):
    cell.value = text
    cell.font  = Font(bold=bold, color=_WHITE, size=size)
    cell.fill  = PatternFill("solid", fgColor=_PURPLE)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _subhdr(cell, text):
    cell.value = text
    cell.font  = Font(bold=True, color=_DARK, size=10)
    cell.fill  = PatternFill("solid", fgColor=_PINK_LIGHT)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _section_title(ws, row, col_span, text):
    cell = ws.cell(row, 1, text)
    cell.font  = Font(bold=True, size=12, color=_PURPLE)
    cell.fill  = PatternFill("solid", fgColor=_PINK_LIGHT)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(f"A{row}:{get_column_letter(col_span)}{row}")
    ws.row_dimensions[row].height = 20


def _thin_border():
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)


class DiagnosticCompiler:
    """
    Collects per-model results and writes a 4-sheet Excel workbook after each update.
    """

    def __init__(self):
        self._models = {}   # (exp, model_name) → {"eval": dict, "preds": list}

    # ── Public API ────────────────────────────────────────────────────────────

    def update(self, exp: str, model_name: str,
               eval_data: dict, pred_records: list) -> None:
        self._models[(exp, model_name)] = {"eval": eval_data, "preds": pred_records}
        self.build_excel()

    def load_data(self):
        self._models = {}
        for path in sorted(glob.glob(os.path.join(config.OUTPUT_DIR, "*_eval.json"))):
            with open(path) as f:
                d = json.load(f)
            _migrate_eval(d, path)
            exp   = d.get("exp",   "?")
            model = d.get("model", "?")
            pred_path = os.path.join(config.OUTPUT_DIR,
                                     f"predictions_{exp}_{model}.json")
            preds = json.load(open(pred_path)) if os.path.exists(pred_path) else []
            self._models[(exp, model)] = {"eval": d, "preds": preds}

    def build_excel(self):
        os.makedirs(os.path.join(config.OUTPUT_DIR, "reports"), exist_ok=True)
        date_str = datetime.now().strftime("%Y%m%d")
        path = os.path.join(config.OUTPUT_DIR, "reports",
                            f"breast_cancer_grading_diagnostic_report_{date_str}.xlsx")

        wb = openpyxl.Workbook()
        self._sheet_synthesis(wb)
        self._sheet_performance(wb)
        self._sheet_confusion(wb)
        self._sheet_details(wb)

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        wb.save(_versioned_path(path))

    # ── Sheet 1 — Synthesis & Conclusion ─────────────────────────────────────

    def _sheet_synthesis(self, wb):
        ws = wb.create_sheet("Synthesis & Conclusion")
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 70

        r = 1
        # Title banner
        ws.merge_cells(f"A{r}:B{r}")
        c = ws.cell(r, 1, "Breast Cancer Grading — Diagnostic Report")
        c.font = Font(bold=True, size=16, color=_WHITE)
        c.fill = PatternFill("solid", fgColor=_PURPLE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 30
        r += 1

        ws.cell(r, 1, "Generated").font = Font(bold=True)
        ws.cell(r, 2, datetime.now().strftime("%Y-%m-%d %H:%M"))
        r += 1
        ws.cell(r, 1, "Device").font = Font(bold=True)
        ws.cell(r, 2, str(config.DEVICE))
        r += 1
        ws.cell(r, 1, "Attention type").font = Font(bold=True)
        ws.cell(r, 2, config.ATTENTION_TYPE or "None")
        r += 1
        ws.cell(r, 1, "Models completed").font = Font(bold=True)
        ws.cell(r, 2, len(self._models))
        r += 2

        # ── Section 1: Executive Summary ──────────────────────────────────────
        _section_title(ws, r, 2, "1. Executive Summary")
        r += 1

        if self._models:
            best_key = max(self._models,
                           key=lambda k: self._models[k]["eval"].get("accuracy", 0))
            best_e   = self._models[best_key]["eval"]
            rows = [
                ("Best model",        f"{best_key[0]}_{best_key[1]}"),
                ("Best test accuracy", f"{best_e.get('accuracy', 0):.4f}"),
                ("Macro Precision",   f"{best_e.get('macro_precision', 0):.4f}"),
                ("Macro Recall",      f"{best_e.get('macro_recall', 0):.4f}"),
                ("Macro F1",          f"{best_e.get('macro_f1', 0):.4f}"),
                ("Activation",        best_e.get("activation", "N/A")),
                ("Attention",         best_e.get("attention",  "N/A")),
                ("Epochs trained",    best_e.get("epochs_trained", "N/A")),
                ("Training time (s)", best_e.get("training_time_seconds", "N/A")),
            ]
            for label, val in rows:
                ws.cell(r, 1, label).font = Font(bold=True)
                ws.cell(r, 2, val)
                r += 1
        else:
            ws.cell(r, 2, "No models completed yet.")
            r += 1
        r += 1

        # ── Section 2: Activation Analysis ────────────────────────────────────
        _section_title(ws, r, 2, "2. Activation Function Analysis  (Exp1 ReLU vs Exp2 GELU)")
        r += 1

        exp1_models = {k[1]: v["eval"] for k, v in self._models.items() if k[0] == "Exp1"}
        exp2_models = {k[1]: v["eval"] for k, v in self._models.items() if k[0] == "Exp2"}
        all_backbones = sorted(set(list(exp1_models) + list(exp2_models)))

        if all_backbones:
            ws.cell(r, 1, "Backbone").font = Font(bold=True)
            ws.cell(r, 2, "Δ Accuracy  (GELU − ReLU)").font = Font(bold=True)
            r += 1
            for bb in all_backbones:
                e1_acc = exp1_models.get(bb, {}).get("accuracy", None)
                e2_acc = exp2_models.get(bb, {}).get("accuracy", None)
                if e1_acc is not None and e2_acc is not None:
                    delta = e2_acc - e1_acc
                    sign  = "+" if delta >= 0 else ""
                    ws.cell(r, 1, bb)
                    c = ws.cell(r, 2, f"{sign}{delta:.4f}  "
                                      f"(ReLU {e1_acc:.4f} → GELU {e2_acc:.4f})")
                    c.font = Font(color="006400" if delta >= 0 else "CC0000")
                elif e1_acc is not None:
                    ws.cell(r, 1, bb)
                    ws.cell(r, 2, f"ReLU {e1_acc:.4f}  (GELU not yet run)")
                elif e2_acc is not None:
                    ws.cell(r, 1, bb)
                    ws.cell(r, 2, f"GELU {e2_acc:.4f}  (ReLU not yet run)")
                r += 1
        else:
            ws.cell(r, 2, "Exp1 / Exp2 results not available yet.")
            r += 1
        r += 1

        # ── Section 3: Attention Analysis ─────────────────────────────────────
        _section_title(ws, r, 2,
                       f"3. Attention Analysis  (Exp2 GELU vs Exp3 GELU + {config.ATTENTION_TYPE or 'Attention'})")
        r += 1

        exp3_models = {k[1]: v["eval"] for k, v in self._models.items() if k[0] == "Exp3"}
        all_backbones_23 = sorted(set(list(exp2_models) + list(exp3_models)))

        if all_backbones_23:
            ws.cell(r, 1, "Backbone").font = Font(bold=True)
            ws.cell(r, 2, "Δ Accuracy  (+ Attention)").font = Font(bold=True)
            r += 1
            for bb in all_backbones_23:
                e2_acc = exp2_models.get(bb, {}).get("accuracy", None)
                e3_acc = exp3_models.get(bb, {}).get("accuracy", None)
                if e2_acc is not None and e3_acc is not None:
                    delta = e3_acc - e2_acc
                    sign  = "+" if delta >= 0 else ""
                    ws.cell(r, 1, bb)
                    c = ws.cell(r, 2, f"{sign}{delta:.4f}  "
                                      f"(no-attn {e2_acc:.4f} → attn {e3_acc:.4f})")
                    c.font = Font(color="006400" if delta >= 0 else "CC0000")
                elif e2_acc is not None:
                    ws.cell(r, 1, bb); ws.cell(r, 2, f"GELU {e2_acc:.4f}  (Exp3 not yet run)")
                elif e3_acc is not None:
                    ws.cell(r, 1, bb); ws.cell(r, 2, f"Attn {e3_acc:.4f}  (Exp2 not yet run)")
                r += 1
        else:
            ws.cell(r, 2, "Exp2 / Exp3 results not available yet.")
            r += 1
        r += 1

        # ── Section 4: Fusion Analysis ─────────────────────────────────────────
        _section_title(ws, r, 2,
                       "4. Feature Fusion Analysis  (Exp3 single backbone vs Exp4 3-backbone fusion)")
        r += 1

        exp4_models = {k[1]: v["eval"] for k, v in self._models.items() if k[0] == "Exp4"}
        if exp4_models:
            best_exp3_acc = max((v["eval"].get("accuracy", 0)
                                 for k, v in self._models.items() if k[0] == "Exp3"),
                                default=None)
            for name, e4 in exp4_models.items():
                e4_acc = e4.get("accuracy", 0)
                ws.cell(r, 1, name).font = Font(bold=True)
                if best_exp3_acc is not None:
                    delta = e4_acc - best_exp3_acc
                    sign  = "+" if delta >= 0 else ""
                    c = ws.cell(r, 2, f"Fusion {e4_acc:.4f}  "
                                      f"(best Exp3 {best_exp3_acc:.4f}, Δ {sign}{delta:.4f})")
                    c.font = Font(color="006400" if delta >= 0 else "CC0000")
                else:
                    ws.cell(r, 2, f"Fusion {e4_acc:.4f}  (Exp3 not yet run)")
                r += 1
        else:
            ws.cell(r, 2, "Exp4 results not available yet.")
            r += 1
        r += 1

        # ── Section 5: Clinical Proof ──────────────────────────────────────────
        _section_title(ws, r, 2, "5. Clinical Proof & Interpretation")
        r += 1

        notes = [
            ("Critical error",
             "Grade III misclassified as Grade I is the most dangerous error — "
             "under-treatment of aggressive cancer. Review off-diagonal CM[2,0]."),
            ("Dataset split",
             "70% train / 15% val / 15% test with fixed seed 42 — splits are "
             "deterministic and reproducible across all experiments."),
            ("Augmentation",
             "H-flip, V-flip, ±15° rotation, brightness/contrast jitter applied "
             "only during training. Val and test use clean resize + normalize only."),
            ("Backbone init",
             "All backbones use ImageNet pretrained weights. Classifier head is "
             "re-trained from scratch; backbone fine-tuned end-to-end."),
        ]
        for label, text in notes:
            ws.cell(r, 1, label).font = Font(bold=True)
            c = ws.cell(r, 2, text)
            c.alignment = Alignment(wrap_text=True)
            ws.row_dimensions[r].height = 32
            r += 1

    # ── Sheet 2 — Overall Performance ─────────────────────────────────────────

    def _sheet_performance(self, wb):
        ws = wb.create_sheet("Overall Performance")

        headers = [
            "Exp", "Model", "Activation", "Attention",
            "Train Acc", "Val Acc (Best)", "Test Acc",
            "Macro Precision", "Macro Recall", "Macro F1", "Weighted F1",
            "Train Loss", "Val Loss", "Test Loss",
            "Training Time (s)", "Epochs Trained",
        ]
        widths = [8, 12, 12, 12, 12, 14, 10, 16, 14, 10, 12, 12, 10, 10, 16, 14]

        for col, (h, w) in enumerate(zip(headers, widths), 1):
            _hdr(ws.cell(1, col), h)
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[1].height = 32
        ws.freeze_panes = "A2"

        for row, ((exp, model), data) in enumerate(self._models.items(), 2):
            e = data["eval"]
            # Guard per_class_metrics being a list (legacy)
            per = e.get("per_class_metrics", {})
            if not isinstance(per, dict):
                per = {}

            vals = [
                exp, model,
                e.get("activation",           "N/A"),
                e.get("attention",            "N/A"),
                round(e.get("train_acc_final",       0), 4),
                round(e.get("best_val_acc",          0), 4),
                round(e.get("accuracy",              0), 4),
                round(e.get("macro_precision",       0), 4),
                round(e.get("macro_recall",          0), 4),
                round(e.get("macro_f1",              0), 4),
                round(e.get("weighted_f1",           0), 4),
                round(e.get("train_loss_final",      0), 6),
                round(e.get("val_loss_final",        0), 6),
                round(e.get("test_loss",             0), 6),
                e.get("training_time_seconds",       "N/A"),
                e.get("epochs_trained",              "N/A"),
            ]
            fill = PatternFill("solid", fgColor=_GREY if row % 2 == 0 else _WHITE)
            for col, v in enumerate(vals, 1):
                c = ws.cell(row, col, v)
                c.fill = fill
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = _thin_border()

    # ── Sheet 3 — Confusion Matrices ──────────────────────────────────────────

    def _sheet_confusion(self, wb):
        ws = wb.create_sheet("Confusion Matrices")
        ws.column_dimensions["A"].width = 22
        grade_labels = ["Grade I", "Grade II", "Grade III"]
        r = 1

        for (exp, model), data in self._models.items():
            cm = data["eval"].get("confusion_matrix", [])
            if not cm:
                continue

            # Model header
            ws.merge_cells(f"A{r}:E{r}")
            c = ws.cell(r, 1, f"{exp}  ·  {model}")
            c.font  = Font(bold=True, size=12, color=_WHITE)
            c.fill  = PatternFill("solid", fgColor=_PURPLE)
            c.alignment = Alignment(horizontal="center")
            ws.row_dimensions[r].height = 20
            r += 1

            # Sub-header row
            ws.cell(r, 1, "True \\ Predicted").font = Font(bold=True)
            ws.cell(r, 1).fill = PatternFill("solid", fgColor=_PINK_LIGHT)
            for col, g in enumerate(grade_labels, 2):
                c = ws.cell(r, col, g)
                c.font  = Font(bold=True, color=_WHITE)
                c.fill  = PatternFill("solid", fgColor=_PINK_MED)
                c.alignment = Alignment(horizontal="center")
                ws.column_dimensions[get_column_letter(col)].width = 14
            r += 1

            # Data rows
            for i, lbl in enumerate(grade_labels):
                c_lbl = ws.cell(r, 1, lbl)
                c_lbl.font = Font(bold=True)
                c_lbl.fill = PatternFill("solid", fgColor=_PINK_LIGHT)
                row_total = sum(cm[i]) if i < len(cm) else 0
                for j in range(len(grade_labels)):
                    val = cm[i][j] if i < len(cm) and j < len(cm[i]) else 0
                    pct = f"  ({val/row_total:.0%})" if row_total > 0 else ""
                    c = ws.cell(r, j + 2, f"{val}{pct}")
                    c.alignment = Alignment(horizontal="center")
                    # Highlight diagonal (correct predictions)
                    if i == j:
                        c.fill = PatternFill("solid", fgColor="E8F5E9")
                        c.font = Font(bold=True, color="1B5E20")
                    else:
                        # Highlight worst error: Grade III → Grade I
                        if i == 2 and j == 0:
                            c.fill = PatternFill("solid", fgColor="FFEBEE")
                            c.font = Font(color="B71C1C")
                    c.border = _thin_border()
                r += 1

            # Per-class summary row
            ws.cell(r, 1, "Correct / Total").font = Font(bold=True, color=_PURPLE)
            acc = data["eval"].get("accuracy", 0)
            ws.merge_cells(f"B{r}:D{r}")
            c = ws.cell(r, 2, f"Overall Accuracy: {acc:.4f}")
            c.alignment = Alignment(horizontal="center")
            c.font = Font(bold=True)
            r += 3

    # ── Sheet 4 — Detailed Process ────────────────────────────────────────────

    def _sheet_details(self, wb):
        ws = wb.create_sheet("Detailed Process")

        grade_labels = ["Grade I", "Grade II", "Grade III"]
        # Columns: Exp, Model, Image, True Grade, Pred Grade,
        #          then TP/TN/FP/FN for each grade (4×3 = 12 cols)
        base_headers = ["Exp", "Model", "Image", "True Grade", "Pred Grade"]
        ova_headers  = []
        for g in grade_labels:
            for metric in ("TP", "TN", "FP", "FN"):
                ova_headers.append(f"{metric} ({g})")

        all_headers = base_headers + ova_headers
        widths = [8, 10, 28, 12, 12] + [10] * len(ova_headers)

        for col, (h, w) in enumerate(zip(all_headers, widths), 1):
            _hdr(ws.cell(1, col), h)
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[1].height = 36
        ws.freeze_panes = "A2"

        row = 2
        for (exp, model), data in self._models.items():
            preds = data.get("preds", [])
            if not preds:
                continue

            for rec in preds:
                true_idx = rec.get("label", -1)
                pred_idx = rec.get("pred",  -1)
                fname    = rec.get("path",  "—")

                true_name = grade_labels[true_idx] if 0 <= true_idx < 3 else "?"
                pred_name = grade_labels[pred_idx] if 0 <= pred_idx < 3 else "?"

                base_vals = [exp, model, fname, true_name, pred_name]

                ova_vals = []
                for g_idx in range(3):
                    tp = 1 if (true_idx == g_idx and pred_idx == g_idx) else 0
                    tn = 1 if (true_idx != g_idx and pred_idx != g_idx) else 0
                    fp = 1 if (true_idx != g_idx and pred_idx == g_idx) else 0
                    fn = 1 if (true_idx == g_idx and pred_idx != g_idx) else 0
                    ova_vals.extend([tp, tn, fp, fn])

                all_vals = base_vals + ova_vals
                fill = PatternFill("solid",
                                   fgColor=_GREY if row % 2 == 0 else _WHITE)
                correct = (true_idx == pred_idx)
                for col, v in enumerate(all_vals, 1):
                    c = ws.cell(row, col, v)
                    c.fill = fill
                    c.alignment = Alignment(horizontal="center")
                    c.border = _thin_border()

                # Highlight Pred Grade cell: green if correct, red if wrong
                pred_cell = ws.cell(row, 5)
                if correct:
                    pred_cell.font = Font(color="1B5E20", bold=True)
                else:
                    pred_cell.font = Font(color="B71C1C", bold=True)
                    # Extra highlight for critical Grade III→Grade I errors
                    if true_idx == 2 and pred_idx == 0:
                        pred_cell.fill = PatternFill("solid", fgColor="FFEBEE")

                row += 1
